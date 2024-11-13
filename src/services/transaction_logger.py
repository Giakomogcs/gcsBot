import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import os
import logging

# Configuração do logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger()

class TransactionLogger:
    def __init__(self, initial_balance, buffer_size=10):
        self.file_name = 'portfolio_history_detailed.xlsx'
        self.initial_balance = initial_balance
        self.cumulative_profit_loss = 0  # Lucro/prejuízo acumulado
        self.buffer_size = buffer_size  # Tamanho do buffer antes de salvar no Excel
        self.transaction_buffer = []  # Buffer para armazenar transações antes de salvar

        # Cria o arquivo se ele não existir
        if not os.path.exists(self.file_name):
            self.create_initial_file()

    def create_initial_file(self):
        columns = [
            'date_time', 'asset', 'transaction_type', 'quantity', 'price_per_asset',
            'transaction_total', 'reason', 'profit_loss', 'total_balance', 'asset_balance',
            'investment_usd', 'returned_to_cash_usd', 'cumulative_profit_loss',
            'usdt_balance', 'roi_percentage', 'cumulative_profit_loss_percentage', 
            'market_value', 'decision_quality', 'performance'
        ]
        df = pd.DataFrame(columns=columns)
        df.to_excel(self.file_name, index=False)
        logger.info("Arquivo Excel inicial criado com sucesso.")

    def record_transaction(self, transaction, portfolio_manager, market_prices):
        # Obtém o nome do ativo principal (ex: de BTCUSDT extrai BTC)
        base_asset = transaction['asset'].replace('USDT', '')

        # Calcula lucro ou prejuízo com base na transação
        profit_loss, investment, returned_to_cash = self.calculate_profit_loss(transaction, portfolio_manager, base_asset)

        # Atualiza o lucro/prejuízo acumulado
        self.cumulative_profit_loss += profit_loss

        # Atualiza saldo total da carteira com preço de mercado atual
        total_balance = self.calculate_total_balance(portfolio_manager, market_prices)

        # Obtém o saldo atual do ativo base e valor de mercado do ativo
        asset_balance = portfolio_manager.assets.get(base_asset, {}).get('quantity', 0)
        market_value = market_prices.get(base_asset, 0) * asset_balance

        # Calcula ROI e porcentagem de lucro/prejuízo acumulado
        roi_percentage = ((total_balance - self.initial_balance) / self.initial_balance) * 100
        cumulative_profit_loss_percentage = (self.cumulative_profit_loss / self.initial_balance) * 100

        # Determina a qualidade da decisão com base no lucro/prejuízo
        decision_quality = "Good" if profit_loss > 0 else "Bad" if profit_loss < 0 else "Neutral"
        performance = f"Profit of {profit_loss:.2f}" if profit_loss > 0 else f"Loss of {profit_loss:.2f}"

        # Constrói a linha de dados detalhada
        data = {
            'date_time': datetime.now(),
            'asset': transaction['asset'],
            'transaction_type': transaction['type'],
            'quantity': f"{transaction['quantity']:.8f}",
            'price_per_asset': f"{transaction['price']:.2f}",
            'transaction_total': f"{transaction['quantity'] * transaction['price']:.2f}",
            'reason': transaction['reason'],
            'profit_loss': f"{profit_loss:.2f}",
            'total_balance': f"{total_balance:.2f}",
            'asset_balance': f"{asset_balance:.8f}",
            'investment_usd': f"{investment:.2f}",
            'returned_to_cash_usd': f"{returned_to_cash:.2f}",
            'cumulative_profit_loss': f"{self.cumulative_profit_loss:.2f}",
            'usdt_balance': f"{portfolio_manager.cash_balance:.2f}",
            'roi_percentage': f"{roi_percentage:.2f}",
            'cumulative_profit_loss_percentage': f"{cumulative_profit_loss_percentage:.2f}",
            'market_value': f"{market_value:.2f}",
            'decision_quality': decision_quality,
            'performance': performance
        }

        # Adiciona a transação ao buffer
        self.transaction_buffer.append(data)
        logger.info(f"Transação adicionada ao buffer: {data}")

        # Grava no Excel se o buffer atingir o tamanho especificado
        if len(self.transaction_buffer) >= self.buffer_size:
            self.flush_buffer_to_excel()

    def calculate_profit_loss(self, transaction, portfolio_manager, base_asset):
        """Calcula lucro/prejuízo, investimento e valor retornado ao caixa."""
        if transaction['type'] == 'sell':
            avg_cost = portfolio_manager.assets[base_asset]['average_cost']
            profit_loss = (transaction['price'] - avg_cost) * transaction['quantity']
            returned_to_cash = transaction['quantity'] * transaction['price']
            investment = 0  # Nenhum investimento em uma venda
        else:
            profit_loss = 0  # Em uma compra inicial, o lucro/prejuízo não é realizado
            returned_to_cash = 0  # Nada retornado ao saldo em uma compra
            investment = transaction['quantity'] * transaction['price']
        return profit_loss, investment, returned_to_cash

    def calculate_total_balance(self, portfolio_manager, market_prices):
        """Calcula o saldo total da carteira com preços de mercado atualizados."""
        total_balance = portfolio_manager.cash_balance + sum(
            asset['quantity'] * market_prices.get(asset_name, asset['average_cost'])
            for asset_name, asset in portfolio_manager.assets.items()
        )
        return total_balance

    def flush_buffer_to_excel(self):
        """Grava as transações do buffer no arquivo Excel."""
        if not self.transaction_buffer:
            return  # Nada para gravar

        try:
            # Carrega o arquivo e verifica se existe a aba onde vamos escrever
            with pd.ExcelWriter(self.file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                df = pd.DataFrame(self.transaction_buffer)
                workbook = writer.book
                if 'Sheet1' not in workbook.sheetnames:
                    self.create_initial_file()

                sheet = writer.sheets.get('Sheet1')
                start_row = sheet.max_row if sheet else 1
                df.to_excel(writer, index=False, header=False, startrow=start_row)

            # Limpa o buffer após salvar
            self.transaction_buffer.clear()
            logger.info(f"{len(df)} transações registradas no Excel.")
        except Exception as e:
            logger.error(f"Erro ao gravar transações no Excel: {e}")

    def export_to_excel(self):
        """Exporta qualquer transação restante no buffer para o Excel."""
        self.flush_buffer_to_excel()
        logger.info(f"Transações registradas em tempo real no arquivo '{self.file_name}'.")
